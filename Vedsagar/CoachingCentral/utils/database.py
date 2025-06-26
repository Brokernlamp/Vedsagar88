import os
import requests
import pandas as pd
from datetime import datetime, date, timedelta
from typing import Dict, List, Any, Optional
import io
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

class DatabaseManager:
    """
    Database manager for NocoDB integration
    Handles all database operations for the EduCRM system
    """
    
    def __init__(self):
        self.base_url = os.getenv("NOCODB_BASE_URL", "http://localhost:8080")
        self.api_token = os.getenv("NOCODB_API_TOKEN", "")
        self.workspace_id = os.getenv("NOCODB_WORKSPACE_ID", "")
        self.base_id = os.getenv("NOCODB_BASE_ID", "")
        
        # Demo mode - use in-memory data if no database configured
        self.demo_mode = not all([self.api_token, self.workspace_id, self.base_id])
        
        self.headers = {
            "xc-token": self.api_token,
            "Content-Type": "application/json"
        }
        
        # Table names mapping
        self.tables = {
            'categories': 'categories',
            'batches': 'batches',
            'students': 'students',
            'tests': 'tests',
            'test_scores': 'test_scores',
            'payments': 'payments',
            'message_templates': 'message_templates',
            'communication_logs': 'communication_logs',
            'activities': 'activities'
        }
        
        # Initialize demo data if in demo mode
        if self.demo_mode:
            self._init_demo_data()
    
    def _init_demo_data(self):
        """Initialize demo data for testing purposes"""
        from datetime import datetime, date, timedelta
        
        # Demo categories
        self.demo_categories = [
            {"id": 1, "name": "NEET Preparation", "description": "Medical entrance exam preparation", "color": "#4CAF50"},
            {"id": 2, "name": "JEE Main & Advanced", "description": "Engineering entrance exam preparation", "color": "#2196F3"},
            {"id": 3, "name": "UPSC Preparation", "description": "Civil services exam preparation", "color": "#FF9800"}
        ]
        
        # Demo batches
        self.demo_batches = [
            {
                "id": 1, "name": "NEET Morning Batch", "category": "NEET Preparation",
                "start_date": date.today(), "end_date": date.today() + timedelta(days=365),
                "capacity": 30, "fee": 50000, "schedule": "Mon-Fri 9AM-12PM",
                "instructor": "Dr. Sharma", "description": "Comprehensive NEET preparation",
                "whatsapp_group_link": "https://chat.whatsapp.com/demo1", "status": "Active"
            },
            {
                "id": 2, "name": "JEE Advanced Batch", "category": "JEE Main & Advanced",
                "start_date": date.today(), "end_date": date.today() + timedelta(days=300),
                "capacity": 25, "fee": 60000, "schedule": "Mon-Fri 2PM-5PM",
                "instructor": "Prof. Kumar", "description": "Advanced JEE preparation",
                "whatsapp_group_link": "https://chat.whatsapp.com/demo2", "status": "Active"
            },
            {
                "id": 3, "name": "UPSC Foundation", "category": "UPSC Preparation",
                "start_date": date.today() + timedelta(days=30), "end_date": date.today() + timedelta(days=730),
                "capacity": 40, "fee": 75000, "schedule": "Weekends 10AM-4PM",
                "instructor": "Ms. Verma", "description": "UPSC foundation course",
                "whatsapp_group_link": "https://chat.whatsapp.com/demo3", "status": "Upcoming"
            }
        ]
        
        # Demo students
        self.demo_students = [
            {
                "id": 1, "full_name": "Priya Sharma", "parent_phone": "9876543210", "student_phone": "9876543211",
                "email": "priya.sharma@email.com", "address": "123 MG Road, Delhi",
                "date_of_birth": date(2005, 3, 15), "category": "NEET Preparation", "batch": "NEET Morning Batch",
                "batch_id": 1, "total_fee": 50000, "paid_amount": 25000, "discount": 0,
                "fee_due_date": date.today() + timedelta(days=15), "admission_date": date.today() - timedelta(days=30),
                "status": "Active", "notes": "Excellent student"
            },
            {
                "id": 2, "full_name": "Rahul Kumar", "parent_phone": "8765432109", "student_phone": "8765432108",
                "email": "rahul.kumar@email.com", "address": "456 CP, Mumbai",
                "date_of_birth": date(2004, 7, 20), "category": "JEE Main & Advanced", "batch": "JEE Advanced Batch",
                "batch_id": 2, "total_fee": 60000, "paid_amount": 60000, "discount": 5000,
                "fee_due_date": date.today() - timedelta(days=10), "admission_date": date.today() - timedelta(days=45),
                "status": "Active", "notes": "Top performer"
            },
            {
                "id": 3, "full_name": "Anita Singh", "parent_phone": "7654321098", "student_phone": "7654321097",
                "email": "anita.singh@email.com", "address": "789 Civil Lines, Lucknow",
                "date_of_birth": date(2003, 12, 5), "category": "UPSC Preparation", "batch": "UPSC Foundation",
                "batch_id": 3, "total_fee": 75000, "paid_amount": 15000, "discount": 0,
                "fee_due_date": date.today() + timedelta(days=5), "admission_date": date.today() - timedelta(days=10),
                "status": "Active", "notes": "Needs guidance"
            }
        ]
        
        # Demo tests
        self.demo_tests = [
            {
                "id": 1, "name": "NEET Mock Test 1", "subject": "Biology", "date": date.today() - timedelta(days=7),
                "max_marks": 200, "category": "NEET Preparation", "batch": "NEET Morning Batch",
                "batch_id": 1, "description": "First mock test", "status": "Completed"
            },
            {
                "id": 2, "name": "JEE Mathematics Test", "subject": "Mathematics", "date": date.today() - timedelta(days=3),
                "max_marks": 100, "category": "JEE Main & Advanced", "batch": "JEE Advanced Batch",
                "batch_id": 2, "description": "Mathematics assessment", "status": "Completed"
            }
        ]
        
        # Demo test scores
        self.demo_test_scores = [
            {"id": 1, "test_id": 1, "student_id": 1, "marks_obtained": 160, "attendance": "Present", "remarks": "Good performance"},
            {"id": 2, "test_id": 2, "student_id": 2, "marks_obtained": 85, "attendance": "Present", "remarks": "Excellent"}
        ]
        
        # Demo payments
        self.demo_payments = [
            {
                "id": 1, "student_id": 1, "amount": 25000, "payment_method": "UPI",
                "payment_date": date.today() - timedelta(days=30), "transaction_reference": "UPI123456",
                "notes": "First installment", "late_fee": 0, "discount": 0, "status": "Completed"
            },
            {
                "id": 2, "student_id": 2, "amount": 65000, "payment_method": "Bank Transfer",
                "payment_date": date.today() - timedelta(days=45), "transaction_reference": "TXN789012",
                "notes": "Full payment with discount", "late_fee": 0, "discount": 5000, "status": "Completed"
            }
        ]
        
        # Demo message templates
        self.demo_templates = [
            {
                "id": 1, "name": "Fee Reminder", "category": "Fee Reminder", "type": "reminder",
                "content": "Dear {student_name}, your fee payment of ₹{pending_amount} is pending. Please pay by {due_date}.",
                "is_active": True, "usage_count": 5
            },
            {
                "id": 2, "name": "Exam Notice", "category": "Exam Notice", "type": "announcement",
                "content": "Dear {student_name}, your {exam_name} is scheduled on {exam_date}. Please prepare well.",
                "is_active": True, "usage_count": 3
            }
        ]
        
        # Demo activities
        self.demo_activities = [
            {
                "id": 1, "description": "New student Priya Sharma enrolled in NEET batch",
                "timestamp": datetime.now() - timedelta(hours=2), "activity_type": "enrollment"
            },
            {
                "id": 2, "description": "Payment of ₹25,000 received from Rahul Kumar",
                "timestamp": datetime.now() - timedelta(hours=4), "activity_type": "payment"
            },
            {
                "id": 3, "description": "Follow-up scheduled with Anita Singh for UPSC batch",
                "timestamp": datetime.now() - timedelta(hours=6), "activity_type": "communication"
            }
        ]
        
        # Demo communication logs
        self.demo_communication_logs = [
            {
                "id": 1, "timestamp": datetime.now() - timedelta(hours=1),
                "recipient_count": 2, "message_preview": "Fee reminder sent to students",
                "template_used": "Fee Reminder", "activity_type": "whatsapp_message"
            }
        ]
        
        # Calculate pending amounts for students
        for student in self.demo_students:
            student['pending_amount'] = student['total_fee'] - student['paid_amount']
    
    def check_connection(self) -> bool:
        """Check if the database connection is working"""
        if self.demo_mode:
            return True
        try:
            response = requests.get(f"{self.base_url}/api/v1/db/meta/projects", headers=self.headers, timeout=10)
            return response.status_code == 200
        except Exception:
            return False
    
    def _make_request(self, method: str, endpoint: str, data: dict = None) -> Optional[Dict]:
        """Make API request to NocoDB"""
        try:
            url = f"{self.base_url}/api/v1/db/data/{self.workspace_id}/{self.base_id}/{endpoint}"
            
            if method.upper() == 'GET':
                response = requests.get(url, headers=self.headers, params=data, timeout=30)
            elif method.upper() == 'POST':
                response = requests.post(url, headers=self.headers, json=data, timeout=30)
            elif method.upper() == 'PUT':
                response = requests.put(url, headers=self.headers, json=data, timeout=30)
            elif method.upper() == 'DELETE':
                response = requests.delete(url, headers=self.headers, timeout=30)
            else:
                return None
            
            if response.status_code in [200, 201]:
                return response.json()
            else:
                print(f"API Error: {response.status_code} - {response.text}")
                return None
        
        except Exception as e:
            print(f"Database request error: {str(e)}")
            return None
    
    # Category Management
    def get_categories(self) -> pd.DataFrame:
        """Get all categories"""
        if self.demo_mode:
            return pd.DataFrame(self.demo_categories)
        try:
            data = self._make_request('GET', self.tables['categories'])
            if data and 'list' in data:
                return pd.DataFrame(data['list'])
            return pd.DataFrame()
        except Exception:
            return pd.DataFrame()
    
    def add_category(self, category_data: dict) -> bool:
        """Add new category"""
        if self.demo_mode:
            # Add to demo categories
            new_id = max([cat['id'] for cat in self.demo_categories]) + 1 if self.demo_categories else 1
            category_data['id'] = new_id
            self.demo_categories.append(category_data)
            return True
        try:
            result = self._make_request('POST', self.tables['categories'], category_data)
            return result is not None
        except Exception:
            return False
    
    def update_category(self, category_id: int, category_data: dict) -> bool:
        """Update category"""
        try:
            result = self._make_request('PUT', f"{self.tables['categories']}/{category_id}", category_data)
            return result is not None
        except Exception:
            return False
    
    def delete_category(self, category_id: int) -> bool:
        """Delete category"""
        try:
            result = self._make_request('DELETE', f"{self.tables['categories']}/{category_id}")
            return result is not None
        except Exception:
            return False
    
    def get_categories_with_stats(self) -> pd.DataFrame:
        """Get categories with student and batch counts"""
        try:
            # Get categories
            categories = self.get_categories()
            if categories.empty:
                return pd.DataFrame()
            
            # Add statistics for each category
            for idx, category in categories.iterrows():
                # Get student count for category
                students = self.get_students_by_category(category['name'])
                categories.loc[idx, 'student_count'] = len(students)
                
                # Get batch count for category
                batches = self.get_batches_by_category(category['name'])
                categories.loc[idx, 'batch_count'] = len(batches)
            
            return categories
        except Exception:
            return pd.DataFrame()
    
    def get_categories_overview(self) -> pd.DataFrame:
        """Get categories overview for dashboard"""
        try:
            categories = self.get_categories_with_stats()
            if categories.empty:
                return pd.DataFrame()
            
            # Add revenue calculation
            for idx, category in categories.iterrows():
                students = self.get_students_by_category(category['name'])
                if not students.empty:
                    revenue = students['paid_amount'].sum() if 'paid_amount' in students.columns else 0
                    categories.loc[idx, 'revenue'] = revenue
                else:
                    categories.loc[idx, 'revenue'] = 0
            
            return categories
        except Exception:
            return pd.DataFrame()
    
    # Batch Management
    def get_all_batches(self) -> pd.DataFrame:
        """Get all batches"""
        if self.demo_mode:
            return pd.DataFrame(self.demo_batches)
        try:
            data = self._make_request('GET', self.tables['batches'])
            if data and 'list' in data:
                return pd.DataFrame(data['list'])
            return pd.DataFrame()
        except Exception:
            return pd.DataFrame()
    
    def get_batches_by_category(self, category: str) -> pd.DataFrame:
        """Get batches filtered by category"""
        try:
            batches = self.get_all_batches()
            if not batches.empty and 'category' in batches.columns:
                return batches[batches['category'] == category]
            return pd.DataFrame()
        except Exception:
            return pd.DataFrame()
    
    def add_batch(self, batch_data: dict) -> bool:
        """Add new batch"""
        try:
            result = self._make_request('POST', self.tables['batches'], batch_data)
            return result is not None
        except Exception:
            return False
    
    def update_batch(self, batch_id: int, batch_data: dict) -> bool:
        """Update batch"""
        try:
            result = self._make_request('PUT', f"{self.tables['batches']}/{batch_id}", batch_data)
            return result is not None
        except Exception:
            return False
    
    def get_batch_capacity_stats(self) -> pd.DataFrame:
        """Get batch capacity statistics"""
        try:
            batches = self.get_all_batches()
            if batches.empty:
                return pd.DataFrame()
            
            # Add current student count for each batch
            for idx, batch in batches.iterrows():
                students = self.get_students_by_batch(batch['id'])
                batches.loc[idx, 'current_students'] = len(students)
            
            return batches
        except Exception:
            return pd.DataFrame()
    
    def get_batch_student_count(self, batch_id: int) -> int:
        """Get number of students in a batch"""
        try:
            students = self.get_students_by_batch(batch_id)
            return len(students)
        except Exception:
            return 0
    
    def get_batch_student_count_by_name(self, batch_name: str) -> int:
        """Get number of students in a batch by name"""
        try:
            students = self.get_all_students()
            if not students.empty and 'batch' in students.columns:
                return len(students[students['batch'] == batch_name])
            return 0
        except Exception:
            return 0
    
    # Student Management
    def get_all_students(self) -> pd.DataFrame:
        """Get all students"""
        if self.demo_mode:
            return pd.DataFrame(self.demo_students)
        try:
            data = self._make_request('GET', self.tables['students'])
            if data and 'list' in data:
                df = pd.DataFrame(data['list'])
                # Calculate pending amount
                if not df.empty and 'total_fee' in df.columns and 'paid_amount' in df.columns:
                    df['pending_amount'] = df['total_fee'] - df['paid_amount']
                return df
            return pd.DataFrame()
        except Exception:
            return pd.DataFrame()
    
    def add_student(self, student_data: dict) -> bool:
        """Add new student"""
        if self.demo_mode:
            # Add to demo students
            new_id = max([student['id'] for student in self.demo_students]) + 1 if self.demo_students else 1
            student_data['id'] = new_id
            student_data['pending_amount'] = student_data.get('total_fee', 0) - student_data.get('paid_amount', 0)
            self.demo_students.append(student_data)
            # Log activity
            self.log_activity(f"New student {student_data['full_name']} enrolled in {student_data['batch']}")
            return True
        try:
            result = self._make_request('POST', self.tables['students'], student_data)
            if result:
                # Log activity
                self.log_activity(f"New student {student_data['full_name']} enrolled in {student_data['batch']}")
                return True
            return False
        except Exception:
            return False
    
    def update_student(self, student_data: dict) -> bool:
        """Update student information"""
        try:
            student_id = student_data.pop('id')
            result = self._make_request('PUT', f"{self.tables['students']}/{student_id}", student_data)
            return result is not None
        except Exception:
            return False
    
    def delete_student(self, student_id: int) -> bool:
        """Delete student"""
        try:
            result = self._make_request('DELETE', f"{self.tables['students']}/{student_id}")
            return result is not None
        except Exception:
            return False
    
    def search_students(self, search_term: str) -> pd.DataFrame:
        """Search students by name or phone"""
        try:
            students = self.get_all_students()
            if students.empty:
                return pd.DataFrame()
            
            # Filter students based on search term
            mask = (
                students['full_name'].str.contains(search_term, case=False, na=False) |
                students['parent_phone'].str.contains(search_term, na=False) |
                students.get('student_phone', pd.Series()).str.contains(search_term, na=False)
            )
            return students[mask]
        except Exception:
            return pd.DataFrame()
    
    def get_students_by_category(self, category: str) -> pd.DataFrame:
        """Get students by category"""
        try:
            students = self.get_all_students()
            if not students.empty and 'category' in students.columns:
                return students[students['category'] == category]
            return pd.DataFrame()
        except Exception:
            return pd.DataFrame()
    
    def get_students_by_batch(self, batch_id: int) -> pd.DataFrame:
        """Get students by batch ID"""
        try:
            students = self.get_all_students()
            if not students.empty and 'batch_id' in students.columns:
                return students[students['batch_id'] == batch_id]
            # Fallback: try to match by batch name if batch_id not available
            elif not students.empty and 'batch' in students.columns:
                batch_data = self._make_request('GET', f"{self.tables['batches']}/{batch_id}")
                if batch_data and 'name' in batch_data:
                    return students[students['batch'] == batch_data['name']]
            return pd.DataFrame()
        except Exception:
            return pd.DataFrame()
    
    def get_students_by_batch_name(self, batch_name: str) -> pd.DataFrame:
        """Get students by batch name"""
        try:
            students = self.get_all_students()
            if not students.empty and 'batch' in students.columns:
                return students[students['batch'] == batch_name]
            return pd.DataFrame()
        except Exception:
            return pd.DataFrame()
    
    def get_students_filtered(self, category_filter: str, batch_filter: str, fee_status_filter: str) -> pd.DataFrame:
        """Get students with applied filters"""
        try:
            students = self.get_all_students()
            if students.empty:
                return pd.DataFrame()
            
            # Apply category filter
            if category_filter != "All Categories" and 'category' in students.columns:
                students = students[students['category'] == category_filter]
            
            # Apply batch filter
            if batch_filter != "All Batches" and 'batch' in students.columns:
                students = students[students['batch'] == batch_filter]
            
            # Apply fee status filter
            if fee_status_filter != "All" and 'pending_amount' in students.columns:
                if fee_status_filter == "Paid":
                    students = students[students['pending_amount'] <= 0]
                elif fee_status_filter == "Pending":
                    students = students[students['pending_amount'] > 0]
                elif fee_status_filter == "Overdue":
                    # Students with pending fees and overdue
                    today = pd.Timestamp.now().date()
                    if 'fee_due_date' in students.columns:
                        students['fee_due_date'] = pd.to_datetime(students['fee_due_date']).dt.date
                        students = students[
                            (students['pending_amount'] > 0) & 
                            (students['fee_due_date'] < today)
                        ]
            
            return students
        except Exception:
            return pd.DataFrame()
    
    def get_students_with_pending_fees(self) -> pd.DataFrame:
        """Get students with pending fees"""
        try:
            students = self.get_all_students()
            if not students.empty and 'pending_amount' in students.columns:
                return students[students['pending_amount'] > 0]
            return pd.DataFrame()
        except Exception:
            return pd.DataFrame()
    
    def get_students_with_filters(self, category: str, fee_status: str, period: str) -> pd.DataFrame:
        """Get students with multiple filters for communication"""
        try:
            students = self.get_all_students()
            if students.empty:
                return pd.DataFrame()
            
            # Apply category filter
            if category != "All Categories" and 'category' in students.columns:
                students = students[students['category'] == category]
            
            # Apply fee status filter
            if fee_status == "Pending Fees Only" and 'pending_amount' in students.columns:
                students = students[students['pending_amount'] > 0]
            elif fee_status == "Paid Up" and 'pending_amount' in students.columns:
                students = students[students['pending_amount'] <= 0]
            
            # Apply period filter
            if period != "All Time" and 'admission_date' in students.columns:
                students['admission_date'] = pd.to_datetime(students['admission_date'])
                today = pd.Timestamp.now()
                
                if period == "This Month":
                    start_date = today.replace(day=1)
                    students = students[students['admission_date'] >= start_date]
                elif period == "Last 3 Months":
                    start_date = today - pd.DateOffset(months=3)
                    students = students[students['admission_date'] >= start_date]
                elif period == "This Year":
                    start_date = today.replace(month=1, day=1)
                    students = students[students['admission_date'] >= start_date]
            
            return students
        except Exception:
            return pd.DataFrame()
    
    # Performance Tracking
    def create_test(self, test_data: dict) -> Optional[int]:
        """Create a new test"""
        try:
            result = self._make_request('POST', self.tables['tests'], test_data)
            if result and 'id' in result:
                return result['id']
            return None
        except Exception:
            return None
    
    def get_recent_tests(self, limit: int = 20) -> pd.DataFrame:
        """Get recent tests"""
        try:
            data = self._make_request('GET', self.tables['tests'], {'limit': limit, 'sort': '-date'})
            if data and 'list' in data:
                return pd.DataFrame(data['list'])
            return pd.DataFrame()
        except Exception:
            return pd.DataFrame()
    
    def get_test_details(self, test_id: int) -> Optional[dict]:
        """Get test details by ID"""
        try:
            return self._make_request('GET', f"{self.tables['tests']}/{test_id}")
        except Exception:
            return None
    
    def save_test_score(self, score_data: dict) -> bool:
        """Save or update test score"""
        try:
            # Check if score already exists
            existing_data = self._make_request('GET', self.tables['test_scores'], {
                'where': f"test_id={score_data['test_id']}&student_id={score_data['student_id']}"
            })
            
            if existing_data and 'list' in existing_data and existing_data['list']:
                # Update existing score
                score_id = existing_data['list'][0]['id']
                result = self._make_request('PUT', f"{self.tables['test_scores']}/{score_id}", score_data)
            else:
                # Create new score
                result = self._make_request('POST', self.tables['test_scores'], score_data)
            
            return result is not None
        except Exception:
            return False
    
    def get_test_scores(self, test_id: int) -> pd.DataFrame:
        """Get scores for a test"""
        try:
            data = self._make_request('GET', self.tables['test_scores'], {'where': f"test_id={test_id}"})
            if data and 'list' in data:
                return pd.DataFrame(data['list'])
            return pd.DataFrame()
        except Exception:
            return pd.DataFrame()
    
    def get_test_scores_count(self, test_id: int) -> int:
        """Get number of scores recorded for a test"""
        try:
            scores = self.get_test_scores(test_id)
            return len(scores)
        except Exception:
            return 0
    
    def get_student_performance_history(self, student_id: int) -> pd.DataFrame:
        """Get performance history for a student"""
        try:
            # Get test scores for student
            data = self._make_request('GET', self.tables['test_scores'], {'where': f"student_id={student_id}"})
            if data and 'list' in data:
                scores_df = pd.DataFrame(data['list'])
                
                if not scores_df.empty:
                    # Add test details
                    for idx, score in scores_df.iterrows():
                        test_details = self.get_test_details(score['test_id'])
                        if test_details:
                            scores_df.loc[idx, 'test_name'] = test_details['name']
                            scores_df.loc[idx, 'test_date'] = test_details['date']
                            scores_df.loc[idx, 'max_marks'] = test_details['max_marks']
                            scores_df.loc[idx, 'subject'] = test_details.get('subject', '')
                    
                    # Calculate percentage
                    if 'marks_obtained' in scores_df.columns and 'max_marks' in scores_df.columns:
                        scores_df['percentage'] = (scores_df['marks_obtained'] / scores_df['max_marks'] * 100).round(2)
                
                return scores_df
            return pd.DataFrame()
        except Exception:
            return pd.DataFrame()
    
    # Fee Management
    def get_fee_statistics(self) -> Optional[dict]:
        """Get fee collection statistics"""
        try:
            students = self.get_all_students()
            if students.empty:
                return None
            
            total_expected = students['total_fee'].sum() if 'total_fee' in students.columns else 0
            total_collected = students['paid_amount'].sum() if 'paid_amount' in students.columns else 0
            total_pending = students['pending_amount'].sum() if 'pending_amount' in students.columns else 0
            pending_count = len(students[students['pending_amount'] > 0]) if 'pending_amount' in students.columns else 0
            
            collection_rate = (total_collected / total_expected * 100) if total_expected > 0 else 0
            
            # Calculate growth rates (mock data for now)
            return {
                'total_expected': total_expected,
                'total_collected': total_collected,
                'total_pending': total_pending,
                'pending_count': pending_count,
                'collection_rate': collection_rate,
                'collected_this_month': total_collected * 0.3,  # Mock data
                'expected_growth': 15,  # Mock data
                'collection_growth': 12,  # Mock data
            }
        except Exception:
            return None
    
    def record_payment(self, payment_data: dict) -> Optional[dict]:
        """Record a payment"""
        try:
            # Record payment
            payment_result = self._make_request('POST', self.tables['payments'], payment_data)
            
            if payment_result:
                # Update student's paid amount
                student_id = payment_data['student_id']
                amount = payment_data['amount']
                
                # Get current student data
                student_data = self._make_request('GET', f"{self.tables['students']}/{student_id}")
                if student_data:
                    current_paid = student_data.get('paid_amount', 0)
                    new_paid = current_paid + amount
                    
                    # Update student record
                    self._make_request('PUT', f"{self.tables['students']}/{student_id}", {
                        'paid_amount': new_paid
                    })
                    
                    # Log activity
                    self.log_activity(f"Payment of ₹{amount} received from student ID {student_id}")
                
                return {'payment_id': payment_result.get('id')}
            
            return None
        except Exception:
            return None
    
    def generate_payment_receipt(self, payment_id: int) -> Optional[dict]:
        """Generate payment receipt data"""
        try:
            payment_data = self._make_request('GET', f"{self.tables['payments']}/{payment_id}")
            if payment_data:
                # Get student details
                student_data = self._make_request('GET', f"{self.tables['students']}/{payment_data['student_id']}")
                
                return {
                    'payment_id': payment_id,
                    'student_name': student_data.get('full_name', '') if student_data else '',
                    'amount': payment_data['amount'],
                    'date': payment_data['payment_date'],
                    'method': payment_data['payment_method']
                }
            return None
        except Exception:
            return None
    
    def get_recent_payments(self, limit: int = 10) -> pd.DataFrame:
        """Get recent payments"""
        try:
            data = self._make_request('GET', self.tables['payments'], {'limit': limit, 'sort': '-payment_date'})
            if data and 'list' in data:
                payments_df = pd.DataFrame(data['list'])
                
                # Add student names
                if not payments_df.empty:
                    for idx, payment in payments_df.iterrows():
                        student_data = self._make_request('GET', f"{self.tables['students']}/{payment['student_id']}")
                        if student_data:
                            payments_df.loc[idx, 'student_name'] = student_data.get('full_name', '')
                
                return payments_df
            return pd.DataFrame()
        except Exception:
            return pd.DataFrame()
    
    def get_detailed_pending_fees(self) -> pd.DataFrame:
        """Get detailed pending fees information"""
        try:
            students = self.get_students_with_pending_fees()
            if students.empty:
                return pd.DataFrame()
            
            # Calculate days overdue and due
            today = pd.Timestamp.now().date()
            
            if 'fee_due_date' in students.columns:
                students['fee_due_date'] = pd.to_datetime(students['fee_due_date']).dt.date
                students['days_overdue'] = (today - students['fee_due_date']).dt.days
                students['days_until_due'] = (students['fee_due_date'] - today).dt.days
            else:
                students['days_overdue'] = 0
                students['days_until_due'] = 30  # Default 30 days
            
            return students
        except Exception:
            return pd.DataFrame()
    
    def apply_pending_fees_filters(self, pending_fees: pd.DataFrame, fee_range: str, overdue_filter: str, category_filter: str) -> pd.DataFrame:
        """Apply filters to pending fees data"""
        try:
            if pending_fees.empty:
                return pd.DataFrame()
            
            filtered_data = pending_fees.copy()
            
            # Apply fee range filter
            if fee_range != "All Amounts" and 'pending_amount' in filtered_data.columns:
                if fee_range == "< ₹10,000":
                    filtered_data = filtered_data[filtered_data['pending_amount'] < 10000]
                elif fee_range == "₹10,000 - ₹25,000":
                    filtered_data = filtered_data[
                        (filtered_data['pending_amount'] >= 10000) & 
                        (filtered_data['pending_amount'] <= 25000)
                    ]
                elif fee_range == "> ₹25,000":
                    filtered_data = filtered_data[filtered_data['pending_amount'] > 25000]
            
            # Apply overdue filter
            if overdue_filter != "All" and 'days_overdue' in filtered_data.columns:
                if overdue_filter == "Overdue Only":
                    filtered_data = filtered_data[filtered_data['days_overdue'] > 0]
                elif overdue_filter == "Due Soon (7 days)":
                    filtered_data = filtered_data[
                        (filtered_data['days_overdue'] <= 0) & 
                        (filtered_data['days_until_due'] <= 7)
                    ]
                elif overdue_filter == "Due This Month":
                    filtered_data = filtered_data[
                        (filtered_data['days_overdue'] <= 0) & 
                        (filtered_data['days_until_due'] <= 30)
                    ]
            
            # Apply category filter
            if category_filter != "All Categories" and 'category' in filtered_data.columns:
                filtered_data = filtered_data[filtered_data['category'] == category_filter]
            
            return filtered_data
        except Exception:
            return pd.DataFrame()
    
    # Message Templates and Communication
    def get_message_templates(self) -> pd.DataFrame:
        """Get all message templates"""
        try:
            data = self._make_request('GET', self.tables['message_templates'])
            if data and 'list' in data:
                return pd.DataFrame(data['list'])
            return pd.DataFrame()
        except Exception:
            return pd.DataFrame()
    
    def add_message_template(self, template_data: dict) -> bool:
        """Add new message template"""
        try:
            result = self._make_request('POST', self.tables['message_templates'], template_data)
            return result is not None
        except Exception:
            return False
    
    def update_message_template(self, template_id: int, template_data: dict) -> bool:
        """Update message template"""
        try:
            result = self._make_request('PUT', f"{self.tables['message_templates']}/{template_id}", template_data)
            return result is not None
        except Exception:
            return False
    
    def delete_message_template(self, template_id: int) -> bool:
        """Delete message template"""
        try:
            result = self._make_request('DELETE', f"{self.tables['message_templates']}/{template_id}")
            return result is not None
        except Exception:
            return False
    
    def get_fee_reminder_templates(self) -> pd.DataFrame:
        """Get fee reminder templates"""
        try:
            templates = self.get_message_templates()
            if not templates.empty and 'category' in templates.columns:
                return templates[templates['category'] == 'fee_reminder']
            return pd.DataFrame()
        except Exception:
            return pd.DataFrame()
    
    def log_communication_activity(self, activity_data: dict) -> bool:
        """Log communication activity"""
        try:
            log_data = {
                'timestamp': activity_data['timestamp'].isoformat(),
                'recipient_count': len(activity_data['recipients']),
                'message_preview': activity_data['message'][:100],
                'template_used': activity_data.get('template_used'),
                'activity_type': 'whatsapp_message'
            }
            
            result = self._make_request('POST', self.tables['communication_logs'], log_data)
            return result is not None
        except Exception:
            return False
    
    def get_communication_statistics(self) -> Optional[dict]:
        """Get communication statistics"""
        try:
            data = self._make_request('GET', self.tables['communication_logs'])
            if data and 'list' in data:
                logs_df = pd.DataFrame(data['list'])
                
                if not logs_df.empty:
                    logs_df['timestamp'] = pd.to_datetime(logs_df['timestamp'])
                    today = pd.Timestamp.now().date()
                    
                    return {
                        'total': len(logs_df),
                        'today': len(logs_df[logs_df['timestamp'].dt.date == today]),
                        'this_week': len(logs_df[logs_df['timestamp'] >= pd.Timestamp.now() - pd.Timedelta(days=7)]),
                        'this_month': len(logs_df[logs_df['timestamp'] >= pd.Timestamp.now() - pd.Timedelta(days=30)])
                    }
            
            return {'total': 0, 'today': 0, 'this_week': 0, 'this_month': 0}
        except Exception:
            return {'total': 0, 'today': 0, 'this_week': 0, 'this_month': 0}
    
    def get_communication_logs(self) -> pd.DataFrame:
        """Get communication logs"""
        try:
            data = self._make_request('GET', self.tables['communication_logs'], {'sort': '-timestamp'})
            if data and 'list' in data:
                return pd.DataFrame(data['list'])
            return pd.DataFrame()
        except Exception:
            return pd.DataFrame()
    
    def filter_communication_logs(self, logs_df: pd.DataFrame, date_filter: str) -> pd.DataFrame:
        """Filter communication logs by date"""
        try:
            if logs_df.empty:
                return pd.DataFrame()
            
            logs_df['timestamp'] = pd.to_datetime(logs_df['timestamp'])
            
            if date_filter == "Last 7 Days":
                cutoff_date = pd.Timestamp.now() - pd.Timedelta(days=7)
                return logs_df[logs_df['timestamp'] >= cutoff_date]
            elif date_filter == "Last 30 Days":
                cutoff_date = pd.Timestamp.now() - pd.Timedelta(days=30)
                return logs_df[logs_df['timestamp'] >= cutoff_date]
            else:  # All Time
                return logs_df
        except Exception:
            return pd.DataFrame()
    
    # Dashboard and Analytics
    def get_dashboard_metrics(self) -> dict:
        """Get metrics for dashboard"""
        try:
            students = self.get_all_students()
            batches = self.get_all_batches()
            
            total_students = len(students)
            active_batches = len(batches)
            
            # Calculate revenue and pending fees
            monthly_revenue = students['paid_amount'].sum() if not students.empty and 'paid_amount' in students.columns else 0
            pending_fees = students['pending_amount'].sum() if not students.empty and 'pending_amount' in students.columns else 0
            pending_count = len(students[students['pending_amount'] > 0]) if not students.empty and 'pending_amount' in students.columns else 0
            
            return {
                'total_students': total_students,
                'active_batches': active_batches,
                'monthly_revenue': monthly_revenue,
                'pending_fees': pending_fees,
                'pending_count': pending_count,
                'student_growth': 12,  # Mock data
                'batch_growth': 2,     # Mock data
                'revenue_growth': 15   # Mock data
            }
        except Exception:
            return {
                'total_students': 0,
                'active_batches': 0,
                'monthly_revenue': 0,
                'pending_fees': 0,
                'pending_count': 0,
                'student_growth': 0,
                'batch_growth': 0,
                'revenue_growth': 0
            }
    
    def get_category_distribution(self) -> pd.DataFrame:
        """Get student distribution by category"""
        try:
            students = self.get_all_students()
            if not students.empty and 'category' in students.columns:
                distribution = students['category'].value_counts().reset_index()
                distribution.columns = ['category_name', 'student_count']
                return distribution
            return pd.DataFrame()
        except Exception:
            return pd.DataFrame()
    
    def get_monthly_fee_data(self) -> pd.DataFrame:
        """Get monthly fee collection data"""
        try:
            payments = self.get_recent_payments(limit=1000)  # Get more payments for analysis
            if not payments.empty and 'payment_date' in payments.columns:
                payments['payment_date'] = pd.to_datetime(payments['payment_date'])
                payments['month'] = payments['payment_date'].dt.to_period('M').astype(str)
                
                monthly_data = payments.groupby('month')['amount'].sum().reset_index()
                return monthly_data
            return pd.DataFrame()
        except Exception:
            return pd.DataFrame()
    
    def get_recent_activities(self) -> pd.DataFrame:
        """Get recent activities for dashboard"""
        try:
            data = self._make_request('GET', self.tables['activities'], {'limit': 10, 'sort': '-timestamp'})
            if data and 'list' in data:
                activities_df = pd.DataFrame(data['list'])
                
                # Calculate time ago
                if not activities_df.empty and 'timestamp' in activities_df.columns:
                    activities_df['timestamp'] = pd.to_datetime(activities_df['timestamp'])
                    now = pd.Timestamp.now()
                    
                    activities_df['time_ago'] = (now - activities_df['timestamp']).apply(
                        lambda x: f"{x.days} days ago" if x.days > 0 else f"{x.seconds // 3600} hours ago"
                    )
                
                return activities_df
            return pd.DataFrame()
        except Exception:
            return pd.DataFrame()
    
    def get_pending_fees(self) -> pd.DataFrame:
        """Get pending fees for dashboard"""
        return self.get_students_with_pending_fees()
    
    def get_upcoming_batches(self) -> pd.DataFrame:
        """Get upcoming batches"""
        try:
            batches = self.get_all_batches()
            if not batches.empty and 'start_date' in batches.columns:
                batches['start_date'] = pd.to_datetime(batches['start_date']).dt.date
                today = pd.Timestamp.now().date()
                next_week = today + timedelta(days=7)
                
                upcoming = batches[
                    (batches['start_date'] >= today) & 
                    (batches['start_date'] <= next_week)
                ]
                return upcoming
            return pd.DataFrame()
        except Exception:
            return pd.DataFrame()
    
    # Activity Logging
    def log_activity(self, description: str) -> bool:
        """Log system activity"""
        try:
            activity_data = {
                'description': description,
                'timestamp': datetime.now().isoformat(),
                'activity_type': 'system'
            }
            
            result = self._make_request('POST', self.tables['activities'], activity_data)
            return result is not None
        except Exception:
            return False
    
    # Export Functions
    def export_students_to_excel(self) -> bytes:
        """Export students data to Excel"""
        try:
            students = self.get_all_students()
            
            # Create Excel file
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                students.to_excel(writer, sheet_name='Students', index=False)
            
            output.seek(0)
            return output.getvalue()
        except Exception:
            # Return empty Excel file on error
            output = io.BytesIO()
            wb = Workbook()
            wb.save(output)
            output.seek(0)
            return output.getvalue()
    
    def export_batches_to_excel(self) -> bytes:
        """Export batches data to Excel"""
        try:
            batches = self.get_all_batches()
            
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                batches.to_excel(writer, sheet_name='Batches', index=False)
            
            output.seek(0)
            return output.getvalue()
        except Exception:
            output = io.BytesIO()
            wb = Workbook()
            wb.save(output)
            output.seek(0)
            return output.getvalue()
    
    def export_pending_fees_to_excel(self, pending_fees: pd.DataFrame) -> bytes:
        """Export pending fees to Excel"""
        try:
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                pending_fees.to_excel(writer, sheet_name='Pending Fees', index=False)
            
            output.seek(0)
            return output.getvalue()
        except Exception:
            output = io.BytesIO()
            wb = Workbook()
            wb.save(output)
            output.seek(0)
            return output.getvalue()
    
    def export_communication_logs_to_excel(self) -> bytes:
        """Export communication logs to Excel"""
        try:
            logs = self.get_communication_logs()
            
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                logs.to_excel(writer, sheet_name='Communication Logs', index=False)
            
            output.seek(0)
            return output.getvalue()
        except Exception:
            output = io.BytesIO()
            wb = Workbook()
            wb.save(output)
            output.seek(0)
            return output.getvalue()
    
    # Archive Functions
    def archive_completed_batches(self) -> int:
        """Archive completed batches"""
        try:
            batches = self.get_all_batches()
            if batches.empty or 'end_date' not in batches.columns:
                return 0
            
            batches['end_date'] = pd.to_datetime(batches['end_date']).dt.date
            today = pd.Timestamp.now().date()
            
            completed_batches = batches[batches['end_date'] < today]
            
            # Archive logic would go here
            # For now, just return count
            return len(completed_batches)
        except Exception:
            return 0
    
    def refresh_batch_statistics(self) -> bool:
        """Refresh batch statistics"""
        try:
            # This would typically refresh calculated fields
            # For now, just return success
            return True
        except Exception:
            return False
    
    # Additional helper methods for reports and analytics
    def get_kpi_data(self, start_date: date, end_date: date) -> Optional[dict]:
        """Get KPI data for dashboard reports"""
        try:
            # Mock implementation - replace with actual calculations
            return {
                'total_revenue': 245000,
                'revenue_change': 15,
                'new_enrollments': 12,
                'enrollment_change': 8,
                'collection_rate': 87.5,
                'collection_change': 5,
                'avg_performance': 78.2,
                'performance_change': 3
            }
        except Exception:
            return None
    
    def get_revenue_trend_data(self, start_date: date, end_date: date) -> pd.DataFrame:
        """Get revenue trend data"""
        try:
            # Generate date range
            date_range = pd.date_range(start=start_date, end=end_date, freq='D')
            
            # Mock data - replace with actual revenue calculations
            revenue_data = []
            for single_date in date_range:
                revenue_data.append({
                    'date': single_date.date(),
                    'revenue': 5000 + (single_date.day * 100)  # Mock calculation
                })
            
            return pd.DataFrame(revenue_data)
        except Exception:
            return pd.DataFrame()
    
    def get_enrollment_trend_data(self, start_date: date, end_date: date) -> pd.DataFrame:
        """Get enrollment trend data"""
        try:
            students = self.get_all_students()
            if students.empty or 'admission_date' not in students.columns:
                return pd.DataFrame()
            
            students['admission_date'] = pd.to_datetime(students['admission_date']).dt.date
            
            # Filter by date range
            filtered_students = students[
                (students['admission_date'] >= start_date) & 
                (students['admission_date'] <= end_date)
            ]
            
            if filtered_students.empty:
                return pd.DataFrame()
            
            # Group by date and count enrollments
            enrollment_counts = filtered_students.groupby('admission_date').size().reset_index(name='new_enrollments')
            enrollment_counts.rename(columns={'admission_date': 'date'}, inplace=True)
            
            return enrollment_counts
        except Exception:
            return pd.DataFrame()
    
    def get_category_performance_summary(self, start_date: date, end_date: date) -> pd.DataFrame:
        """Get category performance summary"""
        try:
            categories = self.get_categories()
            if categories.empty:
                return pd.DataFrame()
            
            performance_data = []
            
            for _, category in categories.iterrows():
                students = self.get_students_by_category(category['name'])
                
                if not students.empty:
                    # Calculate metrics for each category
                    avg_performance = 75.0  # Mock data
                    collection_rate = 85.0  # Mock data
                    enrollment_rate = 90.0  # Mock data
                    total_revenue = students['paid_amount'].sum() if 'paid_amount' in students.columns else 0
                    
                    performance_data.append({
                        'category': category['name'],
                        'avg_performance': avg_performance,
                        'collection_rate': collection_rate,
                        'enrollment_rate': enrollment_rate,
                        'total_revenue': total_revenue,
                        'student_count': len(students)
                    })
            
            return pd.DataFrame(performance_data)
        except Exception:
            return pd.DataFrame()
    
    # Additional methods for comprehensive reporting
    def run_system_health_check(self) -> Optional[dict]:
        """Run system health check"""
        try:
            return {
                'database': self.check_connection(),
                'api_connection': True,  # Mock data
                'storage': True,         # Mock data
                'performance': True,     # Mock data
                'recommendations': []    # Mock data
            }
        except Exception:
            return None
    
    def generate_bulk_reports(self, selected_reports: List[str], report_format: str) -> Optional[bytes]:
        """Generate bulk reports"""
        try:
            # Mock implementation - would generate actual reports
            output = io.BytesIO()
            
            if report_format == "Excel (.xlsx)":
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    for report in selected_reports:
                        # Create mock data for each report
                        mock_data = pd.DataFrame({'Report': [report], 'Status': ['Generated']})
                        mock_data.to_excel(writer, sheet_name=report[:31], index=False)  # Excel sheet name limit
            
            output.seek(0)
            return output.getvalue()
        except Exception:
            return None
